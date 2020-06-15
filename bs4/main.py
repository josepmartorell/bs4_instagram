from bs4 import BeautifulSoup
from openpyxl.styles import Font
from selenium import webdriver
from openpyxl import Workbook
from time import sleep
import requests
import shutil
import json
import os


class App:
    def __init__(self, auth='/home/jmartorell/Documents/auth.json',
                 target='/home/jmartorell/Documents/target.json',
                 path='/home/jmartorell/Pictures/instagram'):
        self.target = target
        self.path = path
        self.auth = auth
        self.driver = webdriver.Firefox(
            executable_path="/usr/local/bin/geckodriver")
        self.error = False
        self.main_url = 'https://www.instagram.com'
        self.all_images = []
        self.no_of_posts = 0
        self.driver.get(self.main_url)
        sleep(3)
        self.log_in()
        if self.error is False:
            self.sweep_box()
            self.shoot_target()
        if self.error is False:
            self.scroll_down()
        if self.error is False:
            if not os.path.exists(path):
                os.mkdir(path)
            self.download_images()
        sleep(3)
        # self.driver.close()

    def write_spreadsheet(self, images, caption_path):
        print('Writing subtitles to excel ...')
        filepath = os.path.join(caption_path, 'subtitles.xlsx')
        workbook = Workbook()
        workbook.save(filepath)
        worksheet = workbook.active
        # set style sheet
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40000
        format = worksheet.column_dimensions['A']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = worksheet.column_dimensions['B']
        format.font = Font(bold=True, italic=True, name='Arial')
        # write on spreadsheet
        row = 1
        worksheet.cell(row=row, column=1).value = 'Image name'
        worksheet.cell(row=row, column=2).value = 'Subtitle'
        row += 1
        for index, image in enumerate(images):
            filename = 'image_' + str(index) + '.jpg'
            try:
                caption = image['alt']
            except KeyError:
                caption = 'No caption exists'
            worksheet.cell(row=row, column=1).value = filename
            worksheet.cell(row=row, column=2).value = caption
            row += 1
        # save work
        workbook.save(filepath)  # save file
        workbook.close()

    def download_subtitles(self, images):
        subtitles_folder_path = os.path.join(self.path, 'subtitles')
        if not os.path.exists(subtitles_folder_path):
            os.mkdir(subtitles_folder_path)
        self.write_spreadsheet(images, subtitles_folder_path)

    def download_images(self):
        self.all_images = list(set(self.all_images))
        self.download_subtitles(self.all_images)
        print('Recounting total images to download up to', len(self.all_images), '...')
        for index, image in enumerate(self.all_images):
            filename = 'image_' + str(index) + '.jpg'
            image_path = os.path.join(self.path, filename)
            link = image['src']
            print('Downloading image', index, 'with url:')
            print(link)
            if link != '/static/images/web/mobile_nav_type_logo.png/735145cfe0a4.png':
                response = requests.get(link, stream=True)
            try:
                with open(image_path, 'wb') as file:
                    shutil.copyfileobj(response.raw, file)  # source -  destination
            except Exception as e:
                print(e)
                print('Could not download image number ', index)
                print('Image link -->', link)

    def scroll_down(self):
        try:
            no_of_posts = self.driver.find_element_by_xpath('//span[text()=" posts"]').text
            no_of_posts = no_of_posts.replace(' posts', '')
            no_of_posts = str(no_of_posts).replace(',', '')  # 2,156 --> 2156
            self.no_of_posts = int(no_of_posts)
            if self.no_of_posts > 12:
                no_of_scrolls = int(self.no_of_posts / 12)
                print('Automating ' + str(no_of_scrolls) + ' chained scrolls down ...')
                try:
                    for value in range(no_of_scrolls + 3):
                        soup = BeautifulSoup(self.driver.page_source, 'lxml')
                        for image in soup.find_all('img'):
                            self.all_images.append(image)

                        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
                        sleep(2)
                except Exception as e:
                    self.error = True
                    print(e)
                    print('Some error occurred while trying to scroll down')
            sleep(10)
        except Exception as e:
            print('Could not find no of posts while trying to scroll down\n', e)
            self.error = True

    def shoot_target(self):
        try:
            with open(self.target, 'r') as t:
                target_dict = json.loads(t.read())
            target_shoot = target_dict['target'][1]
            print('Shooting target ' + target_shoot + ' ...')
            search_bar = self.driver.find_element_by_xpath('//input[@placeholder="Search"]')
            search_bar.send_keys(target_shoot)
            target_profile_url = self.main_url + '/' + target_shoot + '/'
            self.driver.get(target_profile_url)
            sleep(3)

        except Exception as e:
            self.error = True
            print('Could not find search bar\n', e)

    def sweep_box(self):
        # reload page
        sleep(2)
        self.driver.get(self.driver.current_url)

        try:
            print('Closing popup window ...')
            not_now_btn = self.driver.find_element_by_xpath(
                '//*[text()="Not Now"]')

            not_now_btn.click()
            sleep(1)
        except Exception as e:
            print(e)

    def close_settings_window_if_there(self):
        try:
            self.driver.switch_to.window(self.driver.window_handles[1])
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
        except Exception as e:
            print(e)

    def log_in(self):
        try:
            print('Logging in with username and password ...')
            with open(self.auth, 'r') as a:
                auth_dict = json.loads(a.read())
            user_name_input = self.driver.find_element_by_xpath(
                '//input[@aria-label="Phone number, username, or email"]')
            user_name_input.send_keys(auth_dict['username'])
            sleep(1)

            password_input = self.driver.find_element_by_xpath('//input[@aria-label="Password"]')
            password_input.send_keys(auth_dict['password'])
            self.driver.implicitly_wait(100)

            user_name_input.submit()
            sleep(1)

            self.close_settings_window_if_there()
        except Exception as e:
            print('Some exception occurred while trying to find username or password field\n', e)
            self.error = True


if __name__ == '__main__':
    app = App()
